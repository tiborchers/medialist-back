# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import requests
import re
from openpyxl import Workbook, load_workbook
try:
    wb = load_workbook('newseries.xlsx')
except:
    wb = Workbook()
    print("Nuevo workbook")
input = 'y'
while input !='n':
    ws = wb.active
    url = raw_input("Ingrese la url de imdb: ")
    headers = {"Accept-Language": "en-US,en;q=0.5"}
    page = requests.get(url, headers =headers)
    soup = BeautifulSoup(page.content, 'html.parser')
    title=soup.find('h1').get_text()
    poster = soup.find('div', class_='poster').find('a').find('img')['src']
    try:
        time = re.sub('[\n\t min]', '', soup.find('time').get_text()).split('h')
    except:
        time = ['0']
    episodes = re.sub("[ \n\tepisodes']", '', soup.find('span', class_="bp_sub_heading").get_text())
    if (len(time) <2):
        time.append(time[0])
        time[0] = 0
    if (len(time[1])<1):
        time[1]=0
    time = int(time[0])*60 + int(time[1])
    note = float(soup.find('span', itemprop = 'ratingValue').get_text())
    genresandyear = soup.find('div', class_='subtext').find_all('a')
    genres = genresandyear[:-1]
    year = re.sub('[–]','/', genresandyear[-1].get_text().encode('utf-8'))
    print(year)
    year = re.sub('[ \n\t() TVSeries]','',year).split('/')
    year = filter(None, year)
    print(year)
    row = ws.max_row + 1
    rd = ws.row_dimensions[row]
    rd.height = 140
    ws['A'+str(row)]= '=IMAGE("'+poster+'")'
    ws['B'+str(row)]= title
    ws['C'+str(row)]= year[0]
    try:
        ws['D'+str(row)]= year[1]
    except:
        pass
    ws['E'+str(row)]= episodes
    ws['F'+str(row)]= time
    ws['G'+str(row)]= re.sub('[\n\t ]', '', genres[0].get_text())
    try:
        ws['H'+str(row)]= re.sub('[\n\t ]', '', genres[1].get_text())
    except:
        pass
    try:
        ws['I'+str(row)]= re.sub('[\n\t ]', '', genres[2].get_text())
    except:
        pass
    ws['J'+str(row)]= note
    wb.save('newseries.xlsx')
    input=raw_input("Continuar (y/n)? ")
