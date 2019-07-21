import requests
import json
import xlrd
import pandas as pd
from openpyxl import load_workbook




wb = xlrd.open_workbook(r'hola.xlsx')
sheet = wb.sheet_by_index(1)

for row in range(1237):
    if row < 1:
        continue
    veamos = {
        "image": sheet.cell_value(row, 0).encode("utf-8"),
        "title": sheet.cell_value(row, 1).encode("utf-8"),
        "year" : sheet.cell_value(row, 2),
        "duration": sheet.cell_value(row, 3),
        "rating": sheet.cell_value(row, 7),
        "genres": list(filter(None, [sheet.cell_value(row, 4).encode("utf-8"),
                                     sheet.cell_value(row, 5).encode("utf-8"),
                                     sheet.cell_value(row, 6).encode("utf-8")]))
    }
    print(veamos)
    r = requests.post('http://127.0.0.1:3000/api/documentaries/', json = veamos )
    print(r)

"""wb = load_workbook(r'nuevo.xlsx')
ws = wb["Documentales"]
print(ws.max_row)
for i in range(150):
    if( i < 2):
        continue
    print('A'+str(i))
    ws['A'+str(i)] = ws['A'+str(i)].value[8:-2]
    print(ws['A'+str(i)].value)
wb.save('hola.xlsx')"""
